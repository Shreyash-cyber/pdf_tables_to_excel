"""
Excel writer: converts TableData objects into a styled .xlsx file
using openpyxl. Faithfully reproduces table structure including
merged cells, multi-level headers, and exact numeric values.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.config import (
    EXCEL_DATA_FONT_SIZE,
    EXCEL_DEFAULT_COL_WIDTH,
    EXCEL_HEADER_FONT_SIZE,
    EXCEL_TITLE_FONT_SIZE,
)
from backend.models import ExtractionResult, TableData

logger = logging.getLogger(__name__)

# ── Style constants ───────────────────────────────────────────────────────
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="E8ECF0", end_color="E8ECF0", fill_type="solid")
_TITLE_FONT = Font(name="Calibri", size=EXCEL_TITLE_FONT_SIZE, bold=True, color="1E3A5F")
_HEADER_FONT = Font(name="Calibri", size=EXCEL_HEADER_FONT_SIZE, bold=True)
_DATA_FONT = Font(name="Calibri", size=EXCEL_DATA_FONT_SIZE)
_BOLD_DATA_FONT = Font(name="Calibri", size=EXCEL_DATA_FONT_SIZE, bold=True)
_SECTION_FONT = Font(name="Calibri", size=EXCEL_DATA_FONT_SIZE, bold=True, color="1E3A5F")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


class ExcelWriter:
    """Write extracted tables to a styled Excel workbook."""

    def write(self, result: ExtractionResult, output_path: str | Path) -> Path:
        """
        Write all extracted tables to an Excel file.
        Each table gets its own worksheet.
        Returns the path to the created file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        # Remove the default sheet
        if wb.active:
            wb.remove(wb.active)

        used_names: set[str] = set()
        page_counter: dict[int, int] = {}  # track table count per page

        for idx, table in enumerate(result.tables):
            if table.is_empty:
                continue

            # Page-based naming: "Page 1", "Page 1 (2)", etc.
            pg = table.page_number
            page_counter[pg] = page_counter.get(pg, 0) + 1
            if page_counter[pg] == 1:
                base_name = f"Page {pg}"
            else:
                base_name = f"Page {pg} ({page_counter[pg]})"

            sheet_name = self._unique_sheet_name(base_name, idx, used_names)
            used_names.add(sheet_name)

            ws = wb.create_sheet(title=sheet_name)
            self._write_table(ws, table)

        # If no sheets were created, add a placeholder
        if not wb.sheetnames:
            ws = wb.create_sheet(title="No Tables Found")
            ws["A1"] = "No tables were detected in the uploaded PDF."
            ws["A1"].font = Font(size=12, italic=True, color="999999")

        wb.save(str(output_path))
        logger.info("Excel file written: %s (%d sheets)", output_path, len(wb.sheetnames))
        return output_path

    def _write_table(self, ws, table: TableData) -> None:
        """Write a single table (title + headers + data) to a worksheet."""
        current_row = 1

        # ── Title row ─────────────────────────────────────────────────────
        if table.title:
            title_val = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', table.title)
            ws.cell(row=current_row, column=1, value=title_val)
            ws.cell(row=current_row, column=1).font = _TITLE_FONT
            ws.cell(row=current_row, column=1).alignment = _LEFT

            # Merge title across all columns
            if table.total_cols > 1:
                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row,
                    end_column=table.total_cols,
                )
            current_row += 1  # blank row after title
            current_row += 1

        # ── Header rows ──────────────────────────────────────────────────
        header_start_row = current_row
        for header_row in table.headers:
            for col_idx, cell_value in enumerate(header_row):
                if isinstance(cell_value, str):
                    cell_value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', cell_value)
                cell = ws.cell(row=current_row, column=col_idx + 1, value=cell_value)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.border = _THIN_BORDER
                cell.alignment = _CENTER
            current_row += 1

        # ── Apply merge regions (offset by header_start_row) ──────────────
        safe_merges = self._sanitize_merge_regions(
            table.merge_regions,
            header_rows=len(table.headers),
            total_cols=table.total_cols,
        )
        for merge in safe_merges:
            try:
                start_r = header_start_row + merge.start_row
                start_c = merge.start_col + 1
                end_r = header_start_row + merge.end_row
                end_c = merge.end_col + 1
                
                ws.merge_cells(
                    start_row=start_r,
                    start_column=start_c,
                    end_row=end_r,
                    end_column=end_c,
                )
                
                # Re-apply styles to all cells in the merged region so that openpyxl
                # colors and borders the entire block, not just the top-left cell.
                for row_cells in ws.iter_rows(min_row=start_r, min_col=start_c, max_row=end_r, max_col=end_c):
                    for cell in row_cells:
                        cell.fill = _HEADER_FILL
                        cell.font = _HEADER_FONT
                        cell.border = _THIN_BORDER
                        cell.alignment = _CENTER

            except Exception as exc:
                logger.debug("Merge region skipped: %s", exc)

        # ── Data rows ─────────────────────────────────────────────────────────────
        for data_row in table.rows:
            # Detect total/subtotal and section header rows for bold styling
            first_cell = str(data_row[0]).strip().upper() if data_row else ""
            is_total_row = any(kw in first_cell for kw in (
                'TOTAL', 'SUB - TOTAL', 'SUB-TOTAL', 'SUBTOTAL',
                'SURPLUS', 'DEFICIT', 'AMOUNT AVAILABLE', 'GRAND TOTAL',
            ))
            is_section_header = any(kw in first_cell for kw in (
                'APPROPRIATIONS', 'INCOME FROM INVESTMENTS', 'OTHER INCOME',
                'CHANGE IN VALUATION', 'PREMIUMS EARNED', 'BENEFITS PAID',
            )) and not any(str(c).strip() for c in data_row[2:])  # label-only

            for col_idx, cell_value in enumerate(data_row):
                cell = ws.cell(row=current_row, column=col_idx + 1)
                cell.border = _THIN_BORDER

                if is_section_header:
                    cell.font = _SECTION_FONT
                elif is_total_row:
                    cell.font = _BOLD_DATA_FONT
                    cell.fill = _TOTAL_FILL
                else:
                    cell.font = _DATA_FONT

                # Strip illegal XML control characters before writing
                if isinstance(cell_value, str):
                    cell_value = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', cell_value)

                # Try to write as a number for Excel compatibility
                numeric = self._try_parse_number(cell_value)
                if numeric is not None:
                    cell.value = numeric
                    cell.alignment = _RIGHT
                    # Apply number format based on the original text
                    cell.number_format = self._infer_number_format(cell_value)
                else:
                    cell.value = cell_value
                    cell.alignment = _LEFT

            current_row += 1

        # ── Auto-fit column widths ────────────────────────────────────────
        self._auto_fit_columns(ws, table)

    def _auto_fit_columns(self, ws, table: TableData) -> None:
        """Set column widths based on content length."""
        all_rows = table.headers + table.rows
        for col_idx in range(table.total_cols):
            max_len = EXCEL_DEFAULT_COL_WIDTH
            for row in all_rows:
                if col_idx < len(row):
                    cell_len = len(str(row[col_idx])) + 2
                    max_len = max(max_len, min(cell_len, 50))  # cap at 50

            col_letter = get_column_letter(col_idx + 1)
            ws.column_dimensions[col_letter].width = max_len

    @staticmethod
    def _sanitize_merge_regions(merges, header_rows: int, total_cols: int):
        """
        Keep only valid, in-header, non-overlapping merge regions.
        This prevents overlapping/invalid header merges from corrupting layout.
        """
        valid = []
        seen = set()

        for m in merges:
            if m.start_row < 0 or m.start_col < 0:
                continue
            if m.end_row < m.start_row or m.end_col < m.start_col:
                continue
            if m.end_row >= header_rows or m.end_col >= total_cols:
                continue
            # Skip single-cell "merges"
            if m.start_row == m.end_row and m.start_col == m.end_col:
                continue

            key = (m.start_row, m.start_col, m.end_row, m.end_col)
            if key in seen:
                continue
            seen.add(key)
            valid.append(m)

        # Sort by area desc then top-left asc so larger parent merges win.
        valid.sort(
            key=lambda x: (
                -((x.end_row - x.start_row + 1) * (x.end_col - x.start_col + 1)),
                x.start_row,
                x.start_col,
            )
        )

        accepted = []

        def _overlap(a, b):
            return not (
                a.end_row < b.start_row
                or b.end_row < a.start_row
                or a.end_col < b.start_col
                or b.end_col < a.start_col
            )

        for cand in valid:
            if any(_overlap(cand, ex) for ex in accepted):
                continue
            accepted.append(cand)

        return accepted

    @staticmethod
    def _unique_sheet_name(title: str, index: int, used: set[str]) -> str:
        """Generate a unique, valid Excel sheet name."""
        name = title.strip() if title else f"Table {index + 1}"

        # Remove invalid characters
        for char in ("\\", "/", "*", "?", ":", "[", "]"):
            name = name.replace(char, "")

        # Truncate
        if len(name) > 31:
            name = name[:28] + "..."

        # Ensure uniqueness
        original = name
        counter = 2
        while name in used:
            suffix = f" ({counter})"
            name = original[: 31 - len(suffix)] + suffix
            counter += 1

        return name

    @staticmethod
    def _try_parse_number(text: str) -> float | int | None:
        """Try to parse a string as a number, preserving integer vs float."""
        if not text or not text.strip():
            return None

        cleaned = text.strip()

        # Remove currency symbols
        for sym in ("$", "₹", "€", "£", "¥"):
            cleaned = cleaned.replace(sym, "")

        # Remove commas (thousand separators)
        cleaned = cleaned.replace(",", "").strip()

        # Handle parenthetical negatives: (1,234.56) → -1234.56
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]

        # Handle percentage — store as decimal (e.g. 53% → 0.53) so Excel
        # format "0.00%" displays correctly (Excel multiplies by 100 for display)
        is_percent = False
        if cleaned.endswith("%"):
            cleaned = cleaned[:-1].strip()
            is_percent = True

        # Dash means zero
        if cleaned in ("-", "—", "–", "‐"):
            return 0

        try:
            val = float(cleaned)
            if is_percent:
                return val / 100.0
            # Return as int if it's a whole number and the original didn't have decimals
            if val == int(val) and "." not in text:
                return int(val)
            return val
        except ValueError:
            return None

    @staticmethod
    def _infer_number_format(original_text: str) -> str:
        """Infer an Excel number format from the original text."""
        text = original_text.strip()

        if "%" in text:
            return "0.00%"

        # Check for decimal places
        # Find the numeric part
        cleaned = text.replace(",", "").replace("$", "").replace("₹", "")
        cleaned = cleaned.replace("€", "").replace("£", "").replace("¥", "")
        if cleaned.startswith("("):
            cleaned = cleaned[1:]
        if cleaned.endswith(")"):
            cleaned = cleaned[:-1]

        if "." in cleaned:
            decimal_part = cleaned.split(".")[-1]
            decimal_places = len(decimal_part)
            fmt = "#,##0." + "0" * decimal_places
        else:
            fmt = "#,##0"

        # Add parenthetical negative format if original used it
        if "(" in text and ")" in text:
            fmt = fmt + "_);(" + fmt + ")"

        return fmt